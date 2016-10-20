# -*- coding: utf-8 -*-
'''再生冷却設計ツール
'''
import sys
reload(sys)
import platform
# デフォルトの文字コードを変更する．
sys.setdefaultencoding('utf-8')
import os
import subprocess
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.font_manager
from matplotlib.font_manager import FontProperties
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
import pprint

if 'Windows' == platform.system():
    font_prop = FontProperties(fname=r'C:\WINDOWS\Fonts\MSGothic.ttf')

if 'Darwin' == platform.system(): # for Mac
    font_prop = FontProperties(fname='/Library/Fonts/Osaka.ttf')
    matplotlib.rcParams['font.family'] = font_prop.get_name()
    matplotlib.rcParams['pdf.fonttype'] = 42
    matplotlib.rcParams['savefig.dpi'] = 300
    matplotlib.rcParams['mathtext.default'] = 'regular'

plt.close('all')

class Parameter():
    def __init__(self, file_name = u"再生冷却.xlsx"):
        print('Paramter initialize...')
        book = file_name
        sheet = "Input"
        EXL = pd.ExcelFile(book)
        self.temporary_filename = '00temp'
        self.df = EXL.parse(sheet)
        self.thrust = self.df.data[u"推力"]
        self.press_chamber = self.df.data[u"燃焼室圧力"]
        self.of_ratio = self.df.data[u"O/F"]
        self.type_fuel = self.df.data[u"推進剤（燃料）"]
        self.type_oxidizer = self.df.data[u"推進剤（酸化剤）"]
        self.density_fuel = self.df.data[u"燃料密度"]
        self.density_oxidizer = self.df.data[u"酸化剤密度"]
        self.L_star = self.df.data[u"L*"]
        self.tensile_strength = self.df.data[u"燃焼室強度"]
        self.safety_coef = self.df.data[u"安全率"]
        self.thermal_capacity_fuel = self.df.data[u"燃料熱容量"]
        self.visc_fuel = self.df.data[u"燃料粘性係数"]
        self.thermal_conductivity_fuel = self.df.data[u"燃料熱伝導率"]
        self.thermal_conductivity_metal = self.df.data[u"金属熱伝導率"]
        self.diameter_chamber = self.df.data[u"燃焼室直径"]
        self.path_width = self.df.data[u"再生冷却_溝幅"]
        self.path_height = self.df.data[u"再生冷却_溝高さ"]
        self.path_num = self.df.data[u"再生冷却_溝数"]

    def __repr__(self):
        pp = pprint.PrettyPrinter(indent=4)
        return str(pp.pprint(self.__dict__))

    def run_CEA(self, type_fuel, type_oxidizer,
                      of_ratio, press_chamber,
                      temp_fuel = 298.15,
                      temp_oxidizer = 90.17):
        file_name = self.temporary_filename # 一時的に作られるファイル
        # ---- .inpファイル作り ----
        input_file = file_name + '.inp'
        str = """problem    o/f=%.2f,
            rocket  equilibrium  frozen  nfz=2
          p,bar = %d
          pi/p= %d
        react
          fuel=%s  wt=100  t,k=%.2f
          oxid=%s wt=100  t,k=%.2f
        output transport
        output short
        output trace=1e-5
            plot p pi/p o/f isp ivac aeat cf t
        end
        """ % (of_ratio, press_chamber*10,
               press_chamber*10,
               type_fuel, temp_fuel,
               type_oxidizer, temp_oxidizer)
        f = open(input_file, 'w') # 書き込みモードで開く
        f.write(str) # 引数の文字列をファイルに書き込む
        f.close() # ファイルを閉じる
        # ---- FCEA2コマンド ----
        cmd = 'FCEA2' # FCEA2にはパスを通しておく
        p = subprocess.Popen(cmd, shell=True, stdin=subprocess.PIPE, stdout=subprocess.PIPE)
        if 'Windows' == platform.system():
            p.communicate(file_name + '\n') # Win環境ではコマンド末尾に¥nを付けないとバグる
        if 'Darwin' == platform.system():
            p.communicate(file_name)

    def read_CEA(self):
        col_name = ['c{0:02d}'.format(i) for i in range(20)]
        df = pd.read_table(self.temporary_filename + ".out",
                           skiprows=10, names = col_name,
                           delim_whitespace=True)
        # データ読み込みする
        # np.array("CHAMBER", "THROAT", "EXIT")
        # _e:equilibrium, _f:frozen
        date_name = ["T,","VISC,MILLIPOISE","Cp,","PRANDTL","Ae/At","CSTAR,","CF","Isp,"]
        self.temp_chamber_e = np.array(df[df.c00 == date_name[0]][0:1][['c02','c03','c04']], dtype=np.float64)
        self.temp_chamber_f = np.array(df[df.c00 == date_name[0]][1:2][['c02','c03','c04']], dtype=np.float64)
        self.visc_gas_e = np.array(df[df.c00 == date_name[1]][0:1][['c01','c02','c03']], dtype=np.float64)
        self.visc_gas_f = np.array(df[df.c00 == date_name[1]][1:2][['c01','c02','c03']], dtype=np.float64)
        self.Cp_e = np.array(df[df.c00 == date_name[2]][0:1][['c02','c03','c04']], dtype=np.float64)
        self.Cp_f = np.array(df[df.c00 == date_name[2]][3:4][['c02','c03','c04']], dtype=np.float64)
        self.prandtl_e = np.array(df[df.c00 == date_name[3]][0:1][['c02','c03','c04']], dtype=np.float64)
        self.prandtl_f = np.array(df[df.c00 == date_name[3]][2:3][['c02','c03','c04']], dtype=np.float64)
        self.throat_ratio_e = np.array(df[df.c00 == date_name[4]][0:1][['c01','c02']], dtype=np.float64)[0][1]
        self.throat_ratio_f = np.array(df[df.c00 == date_name[4]][1:2][['c01','c02']], dtype=np.float64)[0][1]
        self.C_star_e = np.array(df[df.c00 == date_name[5]][0:1][['c02','c03']], dtype=np.float64)[0][1]
        self.C_star_f = np.array(df[df.c00 == date_name[5]][1:2][['c02','c03']], dtype=np.float64)[0][1]
        self.Cf_e = np.array(df[df.c00 == date_name[6]][0:1][['c01','c02']], dtype=np.float64)[0][1]
        self.Cf_f = np.array(df[df.c00 == date_name[6]][1:2][['c01','c02']], dtype=np.float64)[0][1]
        self.Isp_e = np.array(df[df.c00 == date_name[7]][0:1][['c02','c03']], dtype=np.float64)[0][1]
        self.Isp_f = np.array(df[df.c00 == date_name[7]][1:2][['c02','c03']], dtype=np.float64)[0][1]

    def is_frozen(self, is_frozen = True):
        if(is_frozen == True):
            print('/// select frozen ///')
            self.temp_chamber = self.temp_chamber_f
            self.visc_gas = self.visc_gas_f
            self.Cp = self.Cp_f
            self.prandtl = self.prandtl_f
            self.throat_ratio = self.throat_ratio_f
            self.C_star = self.C_star_f
            self.Cf = self.Cf_f
            self.Isp = self.Isp_f
        elif(is_frozen == False):
            print('/// select equilibrium ///')
            self.temp_chamber = self.temp_chamber_e
            self.visc_gas = self.visc_gas_e
            self.Cp = self.Cp_e
            self.prandtl = self.prandtl_e
            self.throat_ratio = self.throat_ratio_e
            self.C_star = self.C_star_e
            self.Cf = self.Cf_e
            self.Isp = self.Isp_e
        else:
            print('*** select Error ***')

    def calc_chamber_spec(self):
        # nozzle
        self.At = self.thrust / self.press_chamber * self.Cf
        self.Dt = np.sqrt(4.0 * self.At / np.pi)
        self.Ae = self.At * self.throat_ratio
        self.De = np.sqrt(4.0 * self.Ae / np.pi)

        # chamber
        self.Vc = self.At * self.L_star * 1000
        self.Dc = self.diameter_chamber
        self.Ac = self.Dc ** 2 * np.pi /4.0
        self.Lc = self.Vc / self.Ac

        # chamber_strength
        self.Ct = (self.press_chamber * self.Dc) / \
                  ((2.0 * self.tensile_strength / self.safety_coef) - \
                  1.2 * self.press_chamber)

        # fuel_consumption
        self.mt = self.thrust / self.Isp
        self.mf = self.mt / (self.of_ratio + 1)
        self.mo = self.mt * self.of_ratio / (self.of_ratio + 1)

    def calc_regene(self):
        # gas_heat_stansfer_coef
        bartz = np.zeros(5)
        bartz[0] = 0.026 * (self.Dt/1000) ** 0.2
        bartz[1] = ((self.visc_gas[0][0]/10000) ** 0.2) * self.Cp[0][0] * 1000 / \
                   (self.prandtl[0][0] ** 0.6)
        bartz[2] = (self.press_chamber * 10e6 / self.C_star) ** 0.8
        bartz[3] = (self.Dt / (1.5 * self.Dt / 2)) ** 0.1
        bartz[4] = (self.At / self.Ac) ** 0.9
        self.hg = bartz.prod()
        self.rg = 1.0 / self.hg

        # fuel_heat_stansfer_coef
        self.path_area = self.path_width * self.path_height
        self.hydraulic_diam = 2.0 * self.path_area / \
                              (self.path_width + self.path_height)
        self.vf = self.mf / self.density_fuel / 1000.0 / \
                  (self.path_area / 10e6) / self.path_num
        self.delta_p = 0.5 * self.density_fuel * 1000 * self.vf ** 2
        self.nyuf = self.visc_fuel / self.density_fuel / 1000.0
        self.Re = self.vf * self.hydraulic_diam * 1000 / self.nyuf
        self.Pr = self.visc_fuel * self.thermal_capacity_fuel
        self.Nu = 0.023 * (self.Re ** 0.8) * (self.Pr ** 0.6)
        self.hf = self.Nu * self.thermal_conductivity_fuel / \
                  self.hydraulic_diam / 10e6
        self.rf = 1.0 / self.hf

        # metal_heat_stansfer_coef
        self.hm = self.thermal_conductivity_metal
        self.rm = (self.Ct * 10e-3) / self.hm

        # total_heat_stansfer_coef
        self.rc = 0.000407663;	# carbon heat resistivity
        self.rt = self.rg + self.rf + self.rm + self.rc
        self.Tc = self.temp_chamber[0][0]
        self.Q = (self.Tc - 298) / self.rt
        self.Tcw = self.Tc - self.Q * (self.rg + self.rc)
        self.Tcwc = self.Tc - self.Q * (self.rg + self.rc + self.rm)

        # delta_fuel_temp(
        self.A_ct = np.pi * (self.Dc + 2 * self.Ct) * self.Lc
        self.A_ct = self.A_ct * 1.2
        self.Tf_out = (self.A_ct * 10e-6) * self.Q / (self.vf * self.thermal_capacity_fuel)

    def make_output_dataframe(self):
        odf = pd.DataFrame([[ u'推力', self.thrust, 'N'],
                            [ u'燃焼室圧力', self.press_chamber, 'MPa'],
                            [ u'O/F', self.of_ratio, '-'],
                            [ u'推進剤（燃料）', self.type_fuel, '-'],
                            [ u'推進剤（酸化剤）', self.type_oxidizer, '-'],
                            [ u'燃料密度', self.density_fuel, 'kg/m3'],
                            [ u'酸化剤密度', self.density_oxidizer, 'kg/m3'],
                            [ u'L*', self.L_star, 'm'],
                            [ u'燃焼室強度', self.tensile_strength, 'MPa'],
                            [ u'安全率', self.safety_coef, '-'],
                            [ u'燃料熱容量', self.thermal_capacity_fuel, ''],
                            [ u'燃料粘性係数', self.visc_fuel, ''],
                            [ u'燃料熱伝導率', self.thermal_conductivity_fuel, ''],
                            [ u'金属熱伝導率', self.thermal_conductivity_metal, ''],
                            [ u'燃焼室直径', self.diameter_chamber, 'mm'],
                            [ u'再生冷却_溝幅', self.path_width, 'mm'],
                            [ u'再生冷却_溝高さ', self.path_height, 'mm'],
                            [ u'再生冷却_溝数', self.path_num, ''],
                            [ u'燃焼室温度', self.temp_chamber[0][0], 'K'],
                            [ u'スロート温度', self.temp_chamber[0][1], 'K'],
                            [ u'出口温度', self.temp_chamber[0][2], 'K'],
                            # [ u'ガス粘性係数', self.visc_gas, ''],
                            # [ u'Cp', self.Cp, ''],
                            # [ u'プランドル数', self.prandtl, ''],
                            [ u'スロート出口比', self.throat_ratio, ''],
                            [ u'C*', self.C_star, ''],
                            [ u'Cf', self.Cf, '-'],
                            [ u'Isp', self.Isp, 's'],
                            [ u'At', self.At, ''],
                            [ u'Dt', self.Dt, ''],
                            [ u'Ae', self.Ae, ''],
                            [ u'De', self.De, ''],
                            [ u'Vc', self.Vc, ''],
                            [ u'Dc', self.Dc, ''],
                            [ u'Ac', self.Ac, ''],
                            [ u'Lc', self.Lc, ''],
                            [ u'Ct', self.Ct, ''],
                            [ u'mt', self.mt, ''],
                            [ u'mf', self.mf, ''],
                            [ u'mo', self.mo, ''],
                            [ u'hg', self.hg, ''],
                            [ u'rg', self.rg, ''],
                            [ u'path_area', self.path_area, ''],
                            [ u'hydraulic_diam', self.hydraulic_diam, ''],
                            [ u'vf', self.vf, ''],
                            [ u'delta_p', self.delta_p, ''],
                            [ u'nyuf', self.nyuf, ''],
                            [ u'Re', self.Re, ''],
                            [ u'Pr', self.Pr, ''],
                            [ u'Nu', self.Nu, ''],
                            [ u'hf', self.hf, ''],
                            [ u'rf', self.rf, ''],
                            [ u'hm', self.hm, ''],
                            [ u'rm', self.rm, ''],
                            [ u'rc', self.rc, ''],
                            [ u'rt', self.rt, ''],
                            [ u'Tc', self.Tc, ''],
                            [ u'Q', self.Q, ''],
                            [ u'Tcw', self.Tcw, ''],
                            [ u'Tcwc', self.Tcwc, ''],
                            [ u'A_ct', self.A_ct, ''],
                            [ u'Tf_out', self.Tf_out, '']
                            ], columns = ['data', 'value', 'unit'])
        return odf

    def output_to_excel(self, df, output_filename, output_sheet='hoge'):
        wb = load_workbook(filename=output_filename)
        ws = wb.create_sheet(title=output_sheet)
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        wb.save(filename=output_filename)

if __name__ == "__main__":
    print("Regenetive cooling...")
    filename = u"再生冷却.xlsx"
    p = Parameter(filename)
    p.run_CEA(p.type_fuel, p.type_oxidizer, p.of_ratio, p.press_chamber)
    p.read_CEA()
    p.is_frozen(True)
    p.calc_chamber_spec()
    p.calc_regene()
    df = p.make_output_dataframe()
    p.output_to_excel(df, filename)
    print("finish")
